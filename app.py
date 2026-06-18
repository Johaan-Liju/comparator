import os
import uuid
from flask import Flask, render_template, request, send_file, jsonify
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

from extractor import (
    extract_policy_data, CORE_ADDONS, EXTENDED_ADDONS,
    _pymupdf_lines, _plumber_text, _plumber_tables, _build_kv,
)

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'uploads')
OUTPUT_FOLDER = os.path.join(os.path.dirname(__file__), 'outputs')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)


def build_comparison_excel(policies: list, output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparison"

    DARK_BLUE = "002060"
    GREEN     = "3A7D22"
    RED       = "FF0000"
    WHITE     = "FFFFFF"

    def fill(hex_color):
        return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

    def bfont(color="000000", size=11):
        return Font(bold=True, color=color, size=size)

    thin   = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right  = Alignment(horizontal="right",  vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # ── logo ────────────────────────────────────────────────────────────────
    logo_path = os.path.join(os.path.dirname(__file__), "logo.jpeg")
    if os.path.exists(logo_path):
        logo_img = XLImage(logo_path)
        logo_img.width  = 130
        logo_img.height = 65
        logo_img.anchor = "A1"
        ws.add_image(logo_img)
        ws.row_dimensions[1].height = 50

    first = policies[0]
    n     = len(policies)
    last_col = get_column_letter(2 + n)

    # ── rows 2-6: shared header ──────────────────────────────────────────────
    header_info = [
        ("Product:",              first.get("product", "Motor Insurance"), False),
        ("Vehicle Model:",        first.get("vehicle_model", "N/A"),       True),
        ("Policy Expiring Date:", first.get("policy_expiring_date", "N/A"),True),
        ("Insured Name:",         first.get("insured_name", "N/A"),        True),
        ("Registration Number:",  first.get("registration_number", "N/A"), True),
    ]
    for i, (label, value, red) in enumerate(header_info, start=2):
        b = ws.cell(row=i, column=2, value=label)
        b.font = bfont(); b.alignment = right
        c = ws.cell(row=i, column=3, value=value)
        c.font = bfont(RED if red else "000000")
        c.alignment = left
        if n > 1:
            ws.merge_cells(f"C{i}:{last_col}{i}")

    # ── row 7: insurers ──────────────────────────────────────────────────────
    lb = ws.cell(row=7, column=2, value="Insurer:")
    lb.font = bfont(); lb.alignment = right
    for i, pol in enumerate(policies):
        c = ws.cell(row=7, column=3+i, value=pol.get("insurer", "N/A"))
        c.font = bfont(); c.alignment = center; c.border = border

    # ── rows 8-11: details ───────────────────────────────────────────────────
    for r, label, key in [
        (8,  "Policy Type:",                   "policy_type"),
        (9,  "Plan Coverage:",                 "plan_coverage"),
        (10, "Insured Declared Value (IDV):",  "idv"),
        (11, "Total Premium Payable:",         "total_premium"),
    ]:
        lb = ws.cell(row=r, column=2, value=label)
        lb.font = bfont(); lb.alignment = right
        for i, pol in enumerate(policies):
            c = ws.cell(row=r, column=3+i, value=pol.get(key))
            c.font = bfont(); c.alignment = center; c.border = border

    # ── row 12: ADD-ONS header ───────────────────────────────────────────────
    for col in range(2, 3+n):
        c = ws.cell(row=12, column=col, value="ADD-ONS:")
        c.font = Font(bold=True, color=WHITE)
        c.fill = fill(DARK_BLUE)
        c.alignment = center; c.border = border

    # ── determine which extended add-ons actually appear in any policy ───────
    active_extended = [
        name for name in EXTENDED_ADDONS
        if any(pol.get("addons", {}).get(name) == "Yes" for pol in policies)
    ]
    addon_rows = CORE_ADDONS + active_extended

    # Collect extra (unknown) add-ons from all policies
    all_extras: list[str] = []
    for pol in policies:
        for e in pol.get("extras", []):
            if e not in all_extras:
                all_extras.append(e)

    # ── add-on rows ──────────────────────────────────────────────────────────
    for ai, addon in enumerate(addon_rows):
        r = 13 + ai
        lb = ws.cell(row=r, column=2, value=f"{addon}:")
        lb.font = Font(bold=True, color=WHITE)
        lb.fill = fill(GREEN)
        lb.alignment = right; lb.border = border
        for i, pol in enumerate(policies):
            val = pol.get("addons", {}).get(addon, "No")
            c = ws.cell(row=r, column=3+i, value=val)
            c.font = Font(bold=True, color="1F7A1F" if val == "Yes" else "CC0000")
            c.alignment = center; c.border = border

    # ── unrecognised extras section ──────────────────────────────────────────
    if all_extras:
        next_r = 13 + len(addon_rows)
        hdr = ws.cell(row=next_r, column=2, value="OTHER DETECTED FEATURES:")
        hdr.font = Font(bold=True, color=WHITE)
        hdr.fill = fill("666666")
        hdr.alignment = center
        for col in range(3, 3+n):
            c = ws.cell(row=next_r, column=col, value="")
            c.fill = fill("666666"); c.border = border

        for ei, extra in enumerate(all_extras):
            r = next_r + 1 + ei
            lb = ws.cell(row=r, column=2, value=extra)
            lb.font = Font(italic=True, color="555555")
            lb.alignment = right; lb.border = border
            for i, pol in enumerate(policies):
                pol_extras = pol.get("extras", [])
                val = "Detected" if extra in pol_extras else ""
                c = ws.cell(row=r, column=3+i, value=val)
                c.font = Font(italic=True, color="555555")
                c.alignment = center; c.border = border

    # ── column widths / row heights ──────────────────────────────────────────
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 44
    for i in range(n):
        ws.column_dimensions[get_column_letter(3+i)].width = 22

    ws.row_dimensions[9].height  = 45
    ws.row_dimensions[12].height = 20
    ws.freeze_panes = "C7"
    wb.save(output_path)


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/analyze", methods=["POST"])
def analyze():
    files = request.files.getlist("policies")
    if not files:
        return jsonify({"error": "No files uploaded"}), 400
    if len(files) > 5:
        return jsonify({"error": "Maximum 5 policy files allowed"}), 400

    policies, errors = [], []

    for f in files:
        if not f.filename:
            continue
        fpath = os.path.join(UPLOAD_FOLDER, f"{uuid.uuid4()}_{f.filename}")
        f.save(fpath)
        try:
            data = extract_policy_data(fpath)
            policies.append(data)
        except Exception as e:
            errors.append(f"{f.filename}: {e}")
        finally:
            try:
                os.remove(fpath)
            except Exception:
                pass

    if not policies:
        return jsonify({"error": "Could not extract data from any file.", "details": errors}), 400

    output_name = f"Policy_Comparison_{uuid.uuid4().hex[:8]}.xlsx"
    output_path = os.path.join(OUTPUT_FOLDER, output_name)
    build_comparison_excel(policies, output_path)

    resp = {"file": output_name, "policies_processed": len(policies)}
    if errors:
        resp["warnings"] = errors
    return jsonify(resp)


@app.route("/debug", methods=["POST"])
def debug_pdf():
    f = request.files.get("policy")
    if not f:
        return jsonify({"error": "No file"}), 400
    fpath = os.path.join(UPLOAD_FOLDER, f"{uuid.uuid4()}_{f.filename}")
    f.save(fpath)
    try:
        full_text, lines = _pymupdf_lines(fpath)
        plumber   = _plumber_text(fpath)
        tables    = _plumber_tables(fpath)
        kv        = _build_kv(lines)
        parsed    = extract_policy_data(fpath)
        return jsonify({
            "pymupdf_text":  full_text[:8000],
            "plumber_text":  plumber[:4000],
            "lines_sample":  lines[:80],
            "tables":        tables[:30],
            "kv_pairs":      dict(list(kv.items())[:60]),
            "parsed_result": parsed,
        })
    finally:
        try:
            os.remove(fpath)
        except Exception:
            pass


@app.route("/download/<filename>")
def download(filename):
    safe = os.path.basename(filename)
    path = os.path.join(OUTPUT_FOLDER, safe)
    if not os.path.exists(path):
        return jsonify({"error": "File not found"}), 404
    return send_file(path, as_attachment=True, download_name=safe)


if __name__ == "__main__":
    app.run(debug=True, port=5000)
